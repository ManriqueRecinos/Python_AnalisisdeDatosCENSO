import openpyxl
from tabulate import tabulate

excel_dataframe = openpyxl.load_workbook("db/poblacion.xlsx")

dataframe = excel_dataframe.active
data = []

for row in range(2, dataframe.max_row):
    _row = [row,]
    for col in dataframe.iter_cols(1,dataframe.max_column):
        _row.append(col[row].value)
        
    data.append(_row)
        
headers= ["#","Departamento","Hombres","Mujeres","Total","%"]
headers_align= (("center",)*6)

print(tabulate( data, headers= headers, tablefmt='fancy_grid',  colalign= headers_align ))